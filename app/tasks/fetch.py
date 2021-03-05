#!/usr/bin/env python
#  coding=utf-8
#  Author:  Yoge
#  Time:  2021/3/5

from openpyxl import load_workbook
from .wukong import fetch_product_price_stat


def fetch_task(file_name):
    print(file_name, "in")
    wb = load_workbook(file_name)
    ws = wb.active
    col_len = len(list(ws.columns))
    ws.insert_cols(col_len, 5)

    rows = ws.iter_rows()
    title_row = next(rows)
    title_row[5].value = "价格众数"
    title_row[6].value = "价格平均数"
    title_row[7].value = "最高价格"
    title_row[8].value = "最低价格"
    title_row[9].value = "统计条数"
    for i, row in enumerate(rows):
        if not row:
            break

        product = row[2].value
        if not product:
            break

        if product == "商品名称":
            continue

        print(product)
        try:
            stat = fetch_product_price_stat(product)
            row[5].value = stat["mode_price"]
            row[6].value = stat['mean_price']
            row[7].value = stat['max_price']
            row[8].value = stat['min_price']
            row[9].value = stat['stat_count']
        except:
            import traceback
            traceback.print_exc()
            row[5].value = "系统错误，请联系我"
    wb.save(file_name)
    print("done!" ,file_name)


